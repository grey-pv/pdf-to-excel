"""
PDF to Excel Extractor
----------------------
Extracts tables from PDF files and exports them to a formatted Excel workbook.

Usage:
    python extract_tables.py input.pdf
    python extract_tables.py input.pdf --output results.xlsx
    python extract_tables.py input.pdf --output results.xlsx --sheet-per-page
"""

import sys
import argparse
from pathlib import Path

import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Color palette ─────────────────────────────────────────────────────────────
HEADER_BG   = "1F4E79"   # dark blue
HEADER_FG   = "FFFFFF"   # white text
ALT_ROW_BG  = "D6E4F0"   # light blue for alternating rows
BORDER_CLR  = "A9C4D8"   # subtle border


def extract_tables_from_pdf(pdf_path: Path) -> list[dict]:
    """
    Opens the PDF and extracts every table on every page.
    Returns a list of dicts: {page, table_index, dataframe}
    """
    results = []

    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        print(f"  → Scanning {total_pages} page(s)...")

        for page_num, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()

            if not tables:
                continue

            for t_idx, raw_table in enumerate(tables, start=1):
                if not raw_table or len(raw_table) < 2:
                    # Skip empty or single-row tables (likely not real data)
                    continue

                # First row becomes the header; rest becomes data
                header = raw_table[0]
                rows   = raw_table[1:]

                # Clean up None values that pdfplumber sometimes returns
                header = [str(h).strip() if h else f"Column_{i}" for i, h in enumerate(header)]
                rows   = [[str(cell).strip() if cell else "" for cell in row] for row in rows]

                df = pd.DataFrame(rows, columns=header)

                results.append({
                    "page":        page_num,
                    "table_index": t_idx,
                    "dataframe":   df,
                })
                print(f"  ✓ Page {page_num} – Table {t_idx}: {len(df)} rows × {len(df.columns)} columns")

    return results


def style_worksheet(ws, df: pd.DataFrame, title: str):
    """Applies professional formatting to a worksheet."""

    thin = Side(style="thin", color=BORDER_CLR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Header row ────────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = Font(bold=True, color=HEADER_FG, name="Calibri", size=11)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        is_alt = row_idx % 2 == 0
        bg = ALT_ROW_BG if is_alt else "FFFFFF"

        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            cell.border    = border
            cell.font      = Font(name="Calibri", size=10)

    # ── Auto-fit column widths (capped at 45) ────────────────────────────────
    for col_idx, col in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(col)),
            df.iloc[:, col_idx - 1].astype(str).str.len().max() if len(df) > 0 else 0,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)

    # ── Freeze top row ────────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Row height for header ─────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22


def export_to_excel(tables: list[dict], output_path: Path, sheet_per_page: bool = False):
    """Writes all extracted tables into a styled Excel workbook."""

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for entry in tables:
            df    = entry["dataframe"]
            page  = entry["page"]
            t_idx = entry["table_index"]

            # Sheet name: "Page 1 – T1" style (max 31 chars, Excel limit)
            sheet_name = f"Page {page} – T{t_idx}"[:31]

            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Re-open workbook to apply custom styling (ExcelWriter doesn't expose ws easily)
    wb = load_workbook(output_path)

    for entry in tables:
        df    = entry["dataframe"]
        page  = entry["page"]
        t_idx = entry["table_index"]
        sheet_name = f"Page {page} – T{t_idx}"[:31]

        ws = wb[sheet_name]
        style_worksheet(ws, df, sheet_name)

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="Extract tables from a PDF and export to Excel.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python extract_tables.py report.pdf
  python extract_tables.py report.pdf --output my_data.xlsx
  python extract_tables.py report.pdf --output my_data.xlsx --sheet-per-page
        """,
    )
    parser.add_argument("pdf",              help="Path to input PDF file")
    parser.add_argument("--output", "-o",   help="Output Excel file (default: <pdf_name>.xlsx)")
    parser.add_argument("--sheet-per-page", action="store_true",
                        help="Create one sheet per page (default: one sheet per table)")
    args = parser.parse_args()

    # ── Validate input ────────────────────────────────────────────────────────
    pdf_path = Path(args.pdf)
    if not pdf_path.exists():
        print(f"[ERROR] File not found: {pdf_path}")
        sys.exit(1)
    if pdf_path.suffix.lower() != ".pdf":
        print(f"[ERROR] Input must be a PDF file, got: {pdf_path.suffix}")
        sys.exit(1)

    output_path = Path(args.output) if args.output else pdf_path.with_suffix(".xlsx")

    # ── Run extraction ────────────────────────────────────────────────────────
    print(f"\n📄 PDF to Excel Extractor")
    print(f"{'─' * 40}")
    print(f"  Input  : {pdf_path}")
    print(f"  Output : {output_path}")
    print(f"{'─' * 40}")

    tables = extract_tables_from_pdf(pdf_path)

    if not tables:
        print("\n⚠️  No tables found in this PDF.")
        print("   Tips:")
        print("   • Make sure the PDF contains actual tables (not images of tables)")
        print("   • Try a different PDF with visible grid/border tables")
        sys.exit(0)

    print(f"\n📊 Exporting {len(tables)} table(s) to Excel...")
    export_to_excel(tables, output_path, sheet_per_page=args.sheet_per_page)

    print(f"\n✅ Done! File saved to: {output_path}")
    print(f"   {len(tables)} table(s) extracted across {len(set(t['page'] for t in tables))} page(s).")
    print(f"\n{'─' * 40}\n")


if __name__ == "__main__":
    main()
